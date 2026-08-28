# -*- coding: utf-8 -*-
"""把轉換結果寫成 F 班 xlsx（四區塊 + 底色 + 每日人力統計）。"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .coverage import daily_shift_counts, SHIFTS

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(THIN, THIN, THIN, THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
CEN = Alignment(horizontal="center", vertical="center")
WEEK = "一二三四五六日"

FIRST_DAY_COL = 6
META_COLS = ["序", "帳號", "核章人員", "人員", "班種"]


def _weekday_label(start_wd, day):
    idx = (start_wd - 1 + (day - 1)) % 7
    return WEEK[idx]


def write(converted, n_days, cfg, out_path, month_label="115.07"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_label
    start_wd = int(cfg.setting("週起始星期", 3) or 3)

    ws.cell(1, 1, f"線東大眾護理之家 {month_label} F班表（程式自動產生）").font = Font(bold=True, size=13)

    row = 3
    order = [("護理", "護理人員"), ("台籍照服", "台籍照服員"),
             ("外籍照服", "外籍照服員"), ("社工", "社工/行政"), ("行政", "行政")]
    for block, title in order:
        ppl = [p for p in converted if p["block"] == block]
        if not ppl:
            continue
        row = _write_block(ws, row, block, title, ppl, n_days, start_wd)
        row += 1

    _write_stats(ws, row + 1, converted, n_days, start_wd, cfg)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 6
    for d in range(1, n_days + 1):
        ws.column_dimensions[get_column_letter(FIRST_DAY_COL + d - 1)].width = 4.2
    ws.freeze_panes = "F4"
    wb.save(out_path)
    return out_path


def _write_block(ws, row, block, title, ppl, n_days, start_wd):
    ws.cell(row, 5, title).font = Font(bold=True, color="1F4E79")
    for d in range(1, n_days + 1):
        c = ws.cell(row, FIRST_DAY_COL + d - 1, d)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CEN
    row += 1
    for i, name in enumerate(META_COLS):
        c = ws.cell(row, i + 1, name)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CEN
    for d in range(1, n_days + 1):
        c = ws.cell(row, FIRST_DAY_COL + d - 1, _weekday_label(start_wd, d))
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CEN
    row += 1
    for i, p in enumerate(ppl, 1):
        stamp = p["record_name"] if p["record_name"] != p["name"] else ""
        ws.cell(row, 1, i)
        ws.cell(row, 2, p["account"])
        ws.cell(row, 3, stamp)
        ws.cell(row, 4, p["name"])
        ws.cell(row, 5, p.get("shift_kind", ""))
        for col in range(1, 6):
            ws.cell(row, col).alignment = CEN
            ws.cell(row, col).border = BORDER
        for d in range(1, n_days + 1):
            info = p["days"].get(d, {})
            code = info.get("code")
            cell = ws.cell(row, FIRST_DAY_COL + d - 1)
            if code and code not in ("REST", None):
                cell.value = code
            elif code == "REST":
                cell.value = "休"
            cell.alignment = CEN
            cell.border = BORDER
            cell.font = Font(size=9)
            col = info.get("color")
            if col:
                cell.fill = PatternFill("solid", fgColor=col)
        row += 1
    return row


def _write_stats(ws, row, converted, n_days, start_wd, cfg):
    counts = daily_shift_counts(converted, n_days)
    min_staff = int(cfg.setting("每班最低人力", 7) or 7)
    ws.cell(row, 5, "每日每班人力統計").font = Font(bold=True, color="C00000")
    for d in range(1, n_days + 1):
        c = ws.cell(row, FIRST_DAY_COL + d - 1, d)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CEN
    row += 1
    stat_rows = [
        ("護理白班", "護理", "D白"), ("護理小夜", "護理", "E小夜"), ("護理大夜", "護理", "N大夜"),
        ("台籍照白", "台籍照服", "D白"), ("台籍照小夜", "台籍照服", "E小夜"), ("台籍照大夜", "台籍照服", "N大夜"),
        ("外籍照白", "外籍照服", "D白"), ("外籍照夜", "外籍照服", "N大夜"),
    ]
    for label, block, shift in stat_rows:
        ws.cell(row, 4, label).font = Font(size=9, bold=True)
        for d in range(1, n_days + 1):
            v = counts[d][shift].get(block, 0)
            ws.cell(row, FIRST_DAY_COL + d - 1, v).alignment = CEN
        row += 1
    for shift in SHIFTS:
        ws.cell(row, 4, f"{shift}合計").font = Font(size=9, bold=True)
        for d in range(1, n_days + 1):
            tot = counts[d][shift]["total"]
            c = ws.cell(row, FIRST_DAY_COL + d - 1, tot)
            c.alignment = CEN
            if tot < min_staff:
                c.fill = PatternFill("solid", fgColor="FFC7CE")
                c.font = Font(size=9, bold=True, color="9C0006")
        row += 1
    return row
