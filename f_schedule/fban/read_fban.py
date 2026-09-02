# -*- coding: utf-8 -*-
"""讀「機構原生 F 班 Excel」→ 還原成 converted 資料，供下游文件工作流使用。

機構的班表特性（已對照真實檔 115.09）：
  * 一個活頁簿有多個月份分頁（114.12、115.01…115.09），要依月份挑對分頁。
  * 每個區塊（護理人員／台籍照服員／外籍照服員／社工）各有一列表頭，
    表頭含「帳號」「核章人員」「班種」與『區塊名稱』欄（該欄下方即人員姓名）。
  * 日期欄在表頭上方另有一列 1..N；每格填 班別碼（D4x/Ex/Nx/例/休/國…）。
  * 樓層以『儲存格底色』表示，且常用 Excel 佈景主題色（theme color），
    需先把 theme 色解析成 RGB 再比對後台〈樓層顏色規則〉。

輸出 converted：[{name, record_name, account, block, shift_kind, is_head, n_days,
                 days:{d:{code,cat,floor,color,is_work}}}]
姓名一律以 record_name（核章人員＝人頭牌照持有人）為主。
"""
import colorsys
import xml.etree.ElementTree as ET
import openpyxl

_A_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
# openpyxl fill 的 theme 整數 → 佈景主題色盤槽位
_THEME_SLOTS = ["dk1", "lt1", "dk2", "lt2", "accent1", "accent2",
                "accent3", "accent4", "accent5", "accent6", "hlink", "folHlink"]

BLOCK_TITLES = {"護理人員": "護理", "台籍照服員": "台籍照服",
                "外籍照服員": "外籍照服", "社工": "社工",
                "社工/行政": "社工", "行政": "行政", "人員": "護理"}
# 本系統 writer 版面：區塊短標題寫在表頭列上一列的第5欄
SHORT_TITLES = {"護理": "護理", "台籍照服": "台籍照服", "外籍照服": "外籍照服",
                "社工": "社工", "社工/行政": "社工", "行政": "行政"}
# 班種被視為「護理長／管理職」者：同日同樓層有多位白班時，不作為責任護士核章。
HEAD_KINDS = {"D0"}
REST_CODES = {"例", "休", "國", "特", "曠", "事", "病"}


# ---------------- 佈景主題色解析 ----------------
def _theme_palette(wb):
    try:
        xml = wb.loaded_theme
        if isinstance(xml, bytes):
            xml = xml.decode("utf-8")
        root = ET.fromstring(xml)
        scheme = root.find(".//a:clrScheme", _A_NS)
        pal = {}
        for child in scheme:
            tag = child.tag.split("}")[1]
            srgb = child.find("a:srgbClr", _A_NS)
            sysc = child.find("a:sysClr", _A_NS)
            if srgb is not None and srgb.get("val"):
                pal[tag] = srgb.get("val")
            elif sysc is not None and sysc.get("lastClr"):
                pal[tag] = sysc.get("lastClr")
        return pal
    except Exception:
        return {}


def _apply_tint(rgb6, tint):
    if not tint:
        return rgb6
    r = int(rgb6[0:2], 16) / 255.0
    g = int(rgb6[2:4], 16) / 255.0
    b = int(rgb6[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint
    l = min(1.0, max(0.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return "%02X%02X%02X" % (round(r * 255), round(g * 255), round(b * 255))


def _cell_rgb6(cell, palette):
    """回傳儲存格底色的 6 碼 RGB（大寫），無填色回 None。"""
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    c = f.fgColor
    if c is None:
        return None
    try:
        if c.type == "rgb" and isinstance(c.rgb, str):
            return c.rgb[-6:].upper()
        if c.type == "theme":
            base = palette.get(_THEME_SLOTS[c.theme])
            if not base:
                return None
            return _apply_tint(base.upper(), float(c.tint or 0)).upper()
    except Exception:
        return None
    return None


def _floor_map(cfg):
    """{RGB6: 樓層}。取後台〈樓層顏色規則〉中有指定樓層者，另補常見夜班色階別名。"""
    m = {}
    for r in cfg.color_rules:
        fl = r.get("樓層")
        rgb = str(r.get("RGB(ARGB)") or "")
        if fl in ("2F", "3F", "5F") and len(rgb) >= 6:
            m.setdefault(rgb[-6:].upper(), fl)
    # 真實班表夜班照服偶用相近色階（機構檔觀察值）：粉紅→3F、中灰→5F
    m.setdefault("FF8AD8", "3F")
    m.setdefault("A5A5A5", "5F")
    return m


# ---------------- 班別碼 → 大類 ----------------
def _cat_of(code):
    if code is None:
        return "空"
    c = str(code).strip()
    if not c or c in REST_CODES:
        return c or "空"
    c0 = c[0].upper()
    if c0 == "D":
        return "D白"
    if c0 == "E":
        return "E小夜"
    if c0 == "N":
        return "N大夜"
    return "其他"


# ---------------- 版面偵測 ----------------
def _pick_sheet(wb, month_label):
    names = wb.sheetnames
    if month_label:
        ml = str(month_label).strip()
        if ml in names:
            return wb[ml]
        for nm in names:  # 例如「115.09」比對「115.09 (2)」
            if nm.strip() == ml:
                return wb[nm]
        for nm in names:
            if ml and ml in nm and "核章" not in nm:
                return wb[nm]
    return wb[names[-1]]  # 預設取最後一張（通常是最新月份）


def _find_day_row(ws):
    """找日期列：某列從某欄起連續 1,2,3…（長度≥20）。回傳 (day_row, day_start_col, n_days)。"""
    for r in range(1, min(ws.max_row, 15) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            if ws.cell(r, c).value == 1:
                k = 1
                while c + k <= ws.max_column and ws.cell(r, c + k).value == k + 1:
                    k += 1
                if k >= 20:
                    return r, c, k
    return None, 6, 31


def _header_cols(ws, r):
    """在表頭列 r 找各欄位置。回傳 dict 或 None（非表頭列）。"""
    cols = {"account": None, "stamp": None, "shift": None,
            "name": None, "block": None}
    for c in range(1, min(ws.max_column, 12) + 1):
        v = ws.cell(r, c).value
        if not isinstance(v, str):
            continue
        v = v.strip()
        if v == "帳號":
            cols["account"] = c
        elif v == "核章人員":
            cols["stamp"] = c
        elif v == "班種":
            cols["shift"] = c
        elif v in BLOCK_TITLES:
            cols["name"] = c
            cols["block"] = BLOCK_TITLES[v]
            cols["generic"] = (v == "人員")
    if cols["stamp"] and cols["name"] and cols["block"]:
        return cols
    return None


def load(xlsx_path, cfg, month_label=None):
    """回傳 (converted, n_days)。"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = _pick_sheet(wb, month_label)
    palette = _theme_palette(wb)
    fmap = _floor_map(cfg)

    day_row, day_start, n_days = _find_day_row(ws)

    # 先掃出所有區塊表頭列
    headers = []
    for r in range(1, ws.max_row + 1):
        hc = _header_cols(ws, r)
        if hc:
            headers.append((r, hc))
    if not headers:
        return [], n_days
    header_rows = [r for r, _ in headers]

    # 統計區起點（本系統版面於此列後為每日人力統計，非人員）
    stats_row = ws.max_row + 1
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 8) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "人力統計" in v:
                stats_row = r
                break
        if stats_row <= ws.max_row:
            break

    converted = []
    for i, (hr, hc) in enumerate(headers):
        block = hc["block"]
        if hc.get("generic") and hr >= 2:  # 本系統版面：由上一列的區塊標題判區塊
            for c in range(1, min(ws.max_column, 12) + 1):
                tv = ws.cell(hr - 1, c).value
                if not isinstance(tv, str):
                    continue
                t = tv.strip()
                # writer 寫的是「護理人員／台籍照服員／外籍照服員」全稱，
                # 手工版面也可能只寫「護理／台籍照服」短稱，兩種都要認。
                if t in BLOCK_TITLES:
                    block = BLOCK_TITLES[t]
                    break
                if t in SHORT_TITLES:
                    block = SHORT_TITLES[t]
                    break
        end = header_rows[i + 1] if i + 1 < len(header_rows) else ws.max_row + 1
        end = min(end, stats_row)
        blanks = 0
        for r in range(hr + 1, end):
            name = ws.cell(r, hc["name"]).value
            if name is None or str(name).strip() == "":
                blanks += 1
                if blanks >= 4:      # 連續空列＝區塊結束（避開下方統計區）
                    break
                continue
            blanks = 0
            name = str(name).strip()
            if name in BLOCK_TITLES or name in ("小計", "合計", "統計"):
                continue
            stamp = ws.cell(r, hc["stamp"]).value if hc["stamp"] else None
            account = ws.cell(r, hc["account"]).value if hc["account"] else None
            shift_kind = ws.cell(r, hc["shift"]).value if hc["shift"] else None
            days = {}
            for d in range(1, n_days + 1):
                cell = ws.cell(r, day_start + d - 1)
                code = cell.value
                if isinstance(code, str):
                    code = code.strip()
                cat = _cat_of(code)
                rgb = _cell_rgb6(cell, palette)
                floor = fmap.get(rgb) if rgb else None
                days[d] = {"code": code, "cat": cat, "floor": floor,
                           "color": rgb,
                           "is_work": cat in ("D白", "E小夜", "N大夜")}
            sk = str(shift_kind).strip() if shift_kind else ""
            converted.append({
                "name": name,           # 名冊欄（護理人員/照服員）＝人頭名，文件顯示用
                "record_name": name,    # 文件一律用名冊欄姓名（非「核章人員」欄）
                "stamp": (str(stamp).strip() if stamp else ""),  # 核章人員欄（僅參考）
                "account": str(account or "").strip(),
                "block": block,
                "shift_kind": sk,
                "is_head": sk.upper() in HEAD_KINDS,
                "n_days": n_days,
                "days": days,
            })
    return converted, n_days
