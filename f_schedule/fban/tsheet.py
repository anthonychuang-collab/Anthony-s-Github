# -*- coding: utf-8 -*-
"""讀取 T 班（xlsx 月分頁），自動偵測日期列與資料列。"""
import openpyxl
from openpyxl.utils import get_column_letter


def _find_day_header(ws, max_scan=15):
    for r in range(1, max_scan + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v == 1:
                n = 1
                cc = c + 1
                while cc <= ws.max_column and ws.cell(r, cc).value == n + 1:
                    n += 1
                    cc += 1
                if n >= 20:
                    return r, c, n
    raise ValueError("找不到日期標題列（應有連續 1..N）")


def _autopick_sheet(wb):
    for name in wb.sheetnames:
        try:
            _find_day_header(wb[name])
            return name
        except ValueError:
            continue
    return wb.sheetnames[-1]


def read(path, sheet=None):
    """回傳 dict：{ 'day_row','first_col','n_days','rows':[{name,account,days:{d:code}}] }
    sheet 留空或找不到時自動挑第一個含日期列的分頁。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    if not sheet or sheet not in wb.sheetnames:
        sheet = _autopick_sheet(wb)
    ws = wb[sheet]
    day_row, first_col, n_days = _find_day_header(ws)
    name_col = first_col - 2
    acct_col = first_col - 1

    people = []
    for r in range(day_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if name is None:
            if all(ws.cell(r + k, name_col).value is None for k in range(0, 4)):
                break
            continue
        name = str(name).strip()
        if not name or name in ("姓名日期", "姓名"):
            continue
        acct = ws.cell(r, acct_col).value
        days = {}
        for d in range(1, n_days + 1):
            v = ws.cell(r, first_col + d - 1).value
            days[d] = None if v is None else str(v).strip()
        people.append({"name": name, "account": acct, "days": days})

    return {"day_row": day_row, "first_col": first_col,
            "n_days": n_days, "rows": people}
