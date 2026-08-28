# -*- coding: utf-8 -*-
"""讀取「後台設定.xlsx」，轉成程式用的主檔物件。"""
import openpyxl
from dataclasses import dataclass, field


@dataclass
class Person:
    name: str
    block: str            # 護理/台籍照服/外籍照服/社工/行政
    account: str = ""     # 帳號 / 外籍代碼
    stamp_name: str = ""  # 核章人員(人頭牌照)；留空代表本人
    floor: str = ""       # 預設樓層 2F/3F/5F
    tw_aide: bool = False  # 是否具台籍照服資格
    shift_kind: str = ""  # 班種
    allowed: set = field(default_factory=set)  # 適用班別碼 {'D4x','D5x','D6x','Ex','Nx','Dx'}
    note: str = ""

    @property
    def record_name(self):
        """護理業務登記時要用的名字（人頭 -> 牌照持有人）。"""
        return self.stamp_name.strip() if self.stamp_name.strip() else self.name


@dataclass
class HeadStaff:
    """人頭池成員。"""
    name: str
    category: str          # 護理 / 台籍照服
    dedicated_to: str = ""  # 專屬對象；留空=通用
    avail_from: tuple = None  # (民國年, 月, 日) 或 None
    avail_to: tuple = None
    note: str = ""

    def available_on(self, roc_year, month, day):
        if self.avail_from and (roc_year, month, day) < self.avail_from:
            return False
        if self.avail_to and (roc_year, month, day) > self.avail_to:
            return False
        return True


@dataclass
class Config:
    settings: dict = field(default_factory=dict)
    people: list = field(default_factory=list)
    code_map: list = field(default_factory=list)   # dict rows
    color_rules: list = field(default_factory=list)  # dict rows
    head_pool: list = field(default_factory=list)   # HeadStaff
    non_head: list = field(default_factory=list)     # 非人頭姓名
    exclude: set = field(default_factory=set)        # 排除人員（不出現在F班）
    month_quota: dict = field(default_factory=dict)  # {(民國年,月): {'實休','例','休','國','節日'}}
    holiday_dates: dict = field(default_factory=dict)  # {(民國年,月,日): 名稱}

    # --- 便利查詢 ---
    def person_by_name(self, name):
        for p in self.people:
            if p.name == str(name).strip():
                return p
        return None

    def people_in(self, block):
        return [p for p in self.people if p.block == block]

    def setting(self, key, default=None):
        return self.settings.get(key, default)

    def quota(self, roc_year, month):
        return self.month_quota.get((roc_year, month))

    def color_for(self, block, category, floor):
        """回傳 (區塊, 班別大類, 樓層) 對應的 ARGB；找不到回 None。"""
        for r in self.color_rules:
            if r["區塊"] != block:
                continue
            if r["班別大類"] != category:
                continue
            fl = r["樓層"]
            if fl == "*" or fl == floor:
                return r["RGB(ARGB)"]
        return None


def _sheet_rows(ws):
    header = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({header[i]: row[i] for i in range(len(header))})
    return out


def _parse_roc(s):
    if not s:
        return None
    parts = str(s).replace("-", ".").replace("/", ".").split(".")
    try:
        return tuple(int(x) for x in parts[:3])
    except ValueError:
        return None


def load(path="後台設定.xlsx"):
    wb = openpyxl.load_workbook(path, data_only=True)
    cfg = Config()

    # 設定
    ws = wb["設定"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        cfg.settings[str(r[0]).strip()] = r[1]

    # 人員主檔
    ws = wb["人員主檔"]
    for d in _sheet_rows(ws):
        if not d.get("姓名"):
            continue
        allowed = set()
        for code in ("D4x", "D5x", "D6x", "Ex", "Nx", "Dx"):
            if str(d.get("可" + code) or "").strip().upper() == "Y":
                allowed.add(code)
        cfg.people.append(Person(
            name=str(d["姓名"]).strip(),
            block=str(d.get("區塊", "")).strip(),
            account=str(d.get("帳號") or "").strip(),
            stamp_name=str(d.get("核章人員(人頭牌照,留空=本人)") or "").strip(),
            floor=str(d.get("預設樓層") or "").strip(),
            tw_aide=str(d.get("台籍照服資格") or "").strip().upper() == "Y",
            shift_kind=str(d.get("班種") or "").strip(),
            allowed=allowed,
            note=str(d.get("備註") or "").strip(),
        ))

    # 班別碼對照 / 樓層顏色規則
    cfg.code_map = _sheet_rows(wb["班別碼對照"])
    cfg.color_rules = _sheet_rows(wb["樓層顏色規則"])

    # 人頭池
    if "人頭池" in wb.sheetnames:
        for d in _sheet_rows(wb["人頭池"]):
            if not d.get("姓名"):
                continue
            cfg.head_pool.append(HeadStaff(
                name=str(d["姓名"]).strip(),
                category=str(d.get("類別", "")).strip(),
                dedicated_to=str(d.get("專屬對象(留空=通用)") or "").strip(),
                avail_from=_parse_roc(d.get("可用起日(民國,留空=不限)")),
                avail_to=_parse_roc(d.get("可用迄日(留空=不限)")),
                note=str(d.get("備註") or "").strip(),
            ))
    if "非人頭其他工作" in wb.sheetnames:
        for d in _sheet_rows(wb["非人頭其他工作"]):
            if d.get("姓名"):
                cfg.non_head.append(str(d["姓名"]).strip())
    if "排除人員" in wb.sheetnames:
        for d in _sheet_rows(wb["排除人員"]):
            if d.get("姓名"):
                cfg.exclude.add(str(d["姓名"]).strip())

    # 年度行事曆（每月例/休/國配額）
    if "年度行事曆" in wb.sheetnames:
        for d in _sheet_rows(wb["年度行事曆"]):
            try:
                y = int(d.get("民國年"))
                m = int(d.get("月"))
            except (TypeError, ValueError):
                continue
            cfg.month_quota[(y, m)] = {
                "實休": int(d.get("機構實休") or 0),
                "例": int(d.get("例") or 0),
                "休": int(d.get("休") or 0),
                "國": int(d.get("國") or 0),
                "節日": str(d.get("國定假日") or "").strip(),
            }
    if "國定假日日期" in wb.sheetnames:
        for d in _sheet_rows(wb["國定假日日期"]):
            date = _parse_roc(d.get("民國年.月.日"))
            if date and len(date) == 3:
                cfg.holiday_dates[date] = str(d.get("名稱") or "").strip()
    return cfg
